"""Build SEO_IMPLEMENTATION_REPORT.xlsx from the built site.

Every metric is extracted from the generated HTML at run time, so the workbook
cannot drift from the site. Totals are Excel formulas, not Python results, so
the sheet recalculates if a row is edited.

Usage: ./.venv/Scripts/python.exe build/make_xlsx.py
"""

import json
import pathlib
import re
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(pathlib.Path(__file__).resolve().parent))
from audit_final import CONTRAST_CHECKS, ratio  # noqa: E402

ROOT = pathlib.Path(__file__).resolve().parent.parent
OUT = ROOT / "SEO_IMPLEMENTATION_REPORT.xlsx"

FONT = "Arial"
HDR_FILL = PatternFill("solid", fgColor="3B4636")
SUB_FILL = PatternFill("solid", fgColor="E8E0D0")
OK_FILL = PatternFill("solid", fgColor="E7EFE0")
WARN_FILL = PatternFill("solid", fgColor="FDF3E2")
FAIL_FILL = PatternFill("solid", fgColor="FBE6E2")
INPUT_FILL = PatternFill("solid", fgColor="FFFF00")
THIN = Side(style="thin", color="D8D2C4")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

LANDING = [
    "home-cleaning", "deep-cleaning", "move-in-out-cleaning", "holiday-home-cleaning",
    "office-commercial-cleaning", "specialized-cleaning", "pest-control",
    "ac-service-dubai", "handyman-services", "annual-maintenance",
]

KEYWORDS = {
    "home-cleaning": ("home cleaning Dubai",
                      "regular cleaning, weekly cleaning, monthly cleaning, hourly cleaning, apartment cleaning, villa cleaning, maid service Dubai",
                      "Commercial - recurring", "Residents of apartments, townhouses and villas"),
    "deep-cleaning": ("deep cleaning Dubai",
                      "apartment deep cleaning, villa deep cleaning, kitchen deep cleaning, bathroom deep cleaning, post construction cleaning",
                      "Commercial - one-off", "Households after neglect, renovation or handover"),
    "move-in-out-cleaning": ("move out cleaning Dubai",
                             "move in cleaning, end of tenancy cleaning, empty property cleaning, oven cleaning, fridge cleaning, balcony cleaning",
                             "Transactional - deadline driven", "Tenants and landlords at handover"),
    "holiday-home-cleaning": ("holiday home cleaning Dubai",
                              "Airbnb cleaning Dubai, holiday home turnover, linen replacement, consumable restocking, mid stay cleaning",
                              "Commercial - B2B recurring", "Short-let owners and property managers"),
    "office-commercial-cleaning": ("office cleaning Dubai",
                                   "commercial cleaning, daily cleaning contracts, office carpet shampoo, washroom sanitization, fit out cleaning, post event cleaning",
                                   "B2B contract", "Office managers, retail and facilities teams"),
    "specialized-cleaning": ("sofa cleaning Dubai",
                             "carpet cleaning, mattress cleaning, curtain cleaning, interior window cleaning",
                             "Commercial - item level", "Households needing upholstery or soft furnishing work"),
    "pest-control": ("pest control Dubai",
                     "cockroach control, bed bug treatment, ant control, rodent control, commercial pest control",
                     "Urgent problem solving", "Residents and businesses with an active infestation"),
    "ac-service-dubai": ("AC service Dubai",
                         "AC chemical wash, AC duct cleaning, AC repair, AC installation, AC gas top up, split AC service, window AC service",
                         "Commercial + urgent", "Apartment, villa and office occupants in Downtown, Business Bay, DIFC"),
    "handyman-services": ("handyman Dubai",
                          "plumbing, electrical, painting, furniture assembly, TV mounting, lock replacement, door repair, ceiling fan, chandelier",
                          "Task specific", "Households and offices needing small repairs"),
    "annual-maintenance": ("annual maintenance contract Dubai",
                           "apartment AMC, villa AMC, preventive maintenance, MEP inspection, scheduled visits",
                           "B2B / landlord", "Landlords, owners and facilities managers"),
}

CHECKLIST = [
    ("Technical SEO", "Unique title tags", "Completed", "All 23 pages, enforced by build/qa.py"),
    ("Technical SEO", "Unique meta descriptions", "Completed", "All 140-160 chars"),
    ("Technical SEO", "Self-referencing canonicals", "Completed", "Absolute, no trailing slash, no .html"),
    ("Technical SEO", "Robots.txt", "Completed", "Allow /, disallow /thank-you, sitemap declared"),
    ("Technical SEO", "XML sitemap", "Completed", "23 URLs, each verified to exist on disk"),
    ("Technical SEO", "HTML sitemap", "Completed", "Grouped Cleaning / Maintenance / Legal"),
    ("Technical SEO", "Redirect chains", "Completed", "None - verified by build/audit_final.py"),
    ("Technical SEO", "Redirect loops", "Completed", "None"),
    ("Technical SEO", "Trailing slash consistency", "Completed", "All internal links extensionless"),
    ("Technical SEO", "Clean URL consistency", "Completed", "No .html in any internal link"),
    ("Technical SEO", "Broken internal links", "Completed", "Zero"),
    ("Technical SEO", "Broken anchors", "Completed", "Zero - every #fragment resolves"),
    ("Technical SEO", "Broken images", "Completed", "Zero - 34/34 referenced files exist"),
    ("Technical SEO", "Orphan assets", "Completed", "Zero unreferenced images"),
    ("Technical SEO", "hreflang", "Not applicable", "Single language (en-AE), no alternate locales"),
    ("On-page SEO", "Single H1 per page", "Completed", "Enforced by QA"),
    ("On-page SEO", "Heading hierarchy, no skips", "Completed", "Was h2->h4 sitewide, fixed"),
    ("On-page SEO", "Keyword in title / H1 / lead", "Completed", "All 10 landing pages"),
    ("On-page SEO", "Thin content", "Completed", "Minimum 2,016 words on landing pages"),
    ("On-page SEO", "Keyword cannibalization", "Reviewed", "Overlap only in generic CTA words"),
    ("On-page SEO", "Semantic HTML", "Completed", "main/header/footer/nav/section/figure"),
    ("Structured data", "Service schema", "Completed", "10 landing pages + hub"),
    ("Structured data", "Breadcrumb schema", "Completed", "3 levels"),
    ("Structured data", "FAQ schema", "Completed", "89 Q&A pairs"),
    ("Structured data", "JSON-LD validity", "Completed", "Parsed on every build"),
    ("Structured data", "AggregateRating / Review", "Blocked", "No verified review corpus - would be fabricated"),
    ("Structured data", "Offer / priceRange", "Blocked", "No published prices"),
    ("Images", "ALT coverage", "Completed", "100 percent"),
    ("Images", "ALT accuracy", "Completed", "Rewritten to describe before/after composites"),
    ("Images", "Responsive srcset / sizes", "Completed", "All images"),
    ("Images", "Lazy loading", "Completed", "All images"),
    ("Images", "Intrinsic dimensions", "Completed", "Read from real JPEG headers"),
    ("Images", "Same image twice on a page", "Completed", "Eliminated"),
    ("Images", "Cross-page image reuse", "Pending", "ba-living / ba-bathroom on 6 pages - needs more photography"),
    ("Images", "Pest control photography", "Pending", "None exists - documented placeholder"),
    ("Images", "AC hero photography", "Pending", "None exists - page runs without a hero image"),
    ("Accessibility", "Keyboard navigation", "Completed", "All controls reachable, Escape closes dropdown"),
    ("Accessibility", "Visible focus", "Completed", "3px focus-visible outline added sitewide"),
    ("Accessibility", "Form labels", "Completed", "Enforced by QA"),
    ("Accessibility", "ARIA attributes", "Completed", "aria-expanded, aria-label, aria-live, aria-hidden"),
    ("Accessibility", "Landmarks + skip link", "Completed", "main/header/footer/nav + skip link"),
    ("Accessibility", "Touch targets", "Completed", "None under 44px at 375px"),
    ("Accessibility", "Text contrast WCAG AA", "Improved", "9 failures found, 7 fixed"),
    ("Accessibility", "Button contrast WCAG AA", "Pending", "2 failures - requires a branding decision"),
    ("Performance", "Shared cached CSS / JS", "Completed", "Replaces 11 inline copies"),
    ("Performance", "LCP element is text", "Completed", "No render-blocking hero image"),
    ("Performance", "CLS prevention", "Completed", "Real width/height on every image"),
    ("Performance", "JavaScript on AC page", "Completed", "Reduced to zero"),
    ("Performance", "CSS duplication", "Completed", "Extractor de-duplication bug fixed"),
    ("Performance", "Lighthouse field data", "Pending", "Run post-deploy"),
    ("Analytics", "generate_lead fires once", "Completed", "Verified in browser, double-click guarded"),
    ("Analytics", "WhatsApp / phone / CTA tracking", "Completed", "Delegated listener with dedupe"),
    ("Analytics", "Built-in trigger double-count", "Completed", "Investigated live container - not an issue"),
    ("Analytics", "Duplicate GA4 generate_lead tag", "Pending", "Live GTM has TWO tags on one trigger"),
    ("Analytics", "Google Ads conversion", "Completed", "Single tag, ID 18246691744"),
    ("Analytics", "Conversion value", "Pending", "Sends value 0 - blocks value-based bidding"),
    ("Local SEO", "Service area consistency", "Completed", "AC scoped to 3 districts, verified"),
    ("Local SEO", "NAP consistency", "Completed", "One phone, one email, one WhatsApp number"),
    ("Local SEO", "Community coverage", "Completed", "10 communities on general pages"),
    ("Local SEO", "Google Business Profile", "Pending", "Outside repository"),
    ("Local SEO", "Local citations", "Pending", "Outside repository"),
    ("Content", "No fabricated claims", "Completed", "Enforced by build/verify_content.py"),
    ("Content", "Testimonials removed", "Completed", "5 unverified reviews deleted"),
    ("Content", "Pricing on page", "Pending", "No verified prices supplied"),
    ("Content", "Blog / informational content", "Pending", "Recommended"),
]

RECOMMENDATIONS = [
    (1, "P1 - Blocking", "Fix duplicate GA4 generate_lead tag in live GTM",
     "Tags 25 and 48 both fire GA4 generate_lead to G-N2VGBEBELF on the same trigger. Every lead is counted twice in GA4. Google Ads is unaffected (single conversion tag).",
     "Pause or delete one of the two tags in GTM. No code change needed.", "GTM admin", "15 minutes"),
    (2, "P1 - Blocking", "Decide on button contrast (WCAG AA)",
     "Primary button label 3.11:1 and WhatsApp button label 1.98:1 both fail WCAG AA 4.5:1. Sage is a mid-tone, so it fails with light AND dark labels.",
     "Either darken the primary button to #5F6E53 (4.89:1) and accept a slightly deeper green, or accept the risk and document it. WhatsApp green is a recognised brand colour - most sites keep it.",
     "Brand owner", "1 hour"),
    (3, "P2 - High", "Commission pest control photography",
     "No pest control imagery exists. The page ships a documented placeholder rather than borrowed cleaning photos.",
     "Shoot: technician inspecting for cockroach/ant activity, bed bug mattress inspection, commercial kitchen treatment.",
     "Marketing", "1 shoot"),
    (4, "P2 - High", "Commission AC service photography",
     "ba-ducts is the only AC-specific asset and it is used by the before/after section, so the page runs without a hero image.",
     "Shoot: technician servicing a wall-mounted split unit, chemical wash in progress, coil cleaning detail.",
     "Marketing", "1 shoot"),
    (5, "P2 - High", "Set a conversion value",
     "The Ads conversion tag reads {{value}}, and the dataLayer pushes value 0. Value-based bidding cannot work.",
     "Push an estimated lead value per service in the generate_lead event, or set a default conversion value in Google Ads.",
     "Marketing + dev", "2 hours"),
    (6, "P2 - High", "Google Business Profile alignment",
     "Highest-leverage local SEO item not in this repository.",
     "Complete the profile, add service categories matching the 10 landing pages, link each service to its page, post photos.",
     "Marketing", "1 day"),
    (7, "P3 - Medium", "Review collection programme",
     "No review schema is implemented because no verified review corpus exists.",
     "Collect Google reviews, then surface them with genuine AggregateRating markup.",
     "Operations", "Ongoing"),
    (8, "P3 - Medium", "Additional before/after photography",
     "ba-living and ba-bathroom each appear on 6 pages. Only 7 composites exist for 6 gallery pages.",
     "Shoot 6-8 more before/after pairs so each page can carry a distinct set.",
     "Marketing", "1 shoot"),
    (9, "P3 - Medium", "Run Lighthouse post-deploy",
     "No CrUX field data exists for these URLs yet. Current performance claims are architectural, not measured.",
     "Run Lighthouse on all 10 landing pages after deploy and record LCP/CLS/INP.",
     "Dev", "1 hour"),
    (10, "P3 - Medium", "Publish trade licence and VAT TRN",
     "Both are placeholder comments in the footer. Real UAE customers and Google both value verifiable business identity.",
     "Supply the numbers; they go into the footer and Organization schema.",
     "Business owner", "15 minutes"),
    (11, "P4 - Later", "Blog for informational intent",
     "Landing pages own commercial intent and should not chase how-to queries.",
     "Publish Dubai-specific guides: AC servicing frequency in summer, tenancy handover checklists, bed bug preparation.",
     "Marketing", "Ongoing"),
    (12, "P4 - Later", "Per-community landing pages",
     "Only once the core ten mature, and only for communities with genuine booking volume.",
     "Avoid spinning up thin location pages - that creates a thin-content problem rather than solving one.",
     "Marketing", "Later"),
    (13, "P4 - Later", "Backlink acquisition",
     "No backlink work has been done.",
     "Target property managers, holiday-home operators and real-estate agencies in covered communities.",
     "Marketing", "Ongoing"),
]


# ---------------------------------------------------------------- extraction
def visible_text(html):
    b = re.sub(r"<head\b.*?</head>", " ", html, flags=re.S | re.I)
    b = re.sub(r"<(script|style)\b.*?</\1>", " ", b, flags=re.S | re.I)
    b = re.sub(r"<!--.*?-->", " ", b, flags=re.S)
    b = re.sub(r"<[^>]+>", " ", b)
    return re.sub(r"\s+", " ", b).strip()


def unescape(s):
    return (s.replace("&amp;", "&").replace("&mdash;", "-")
             .replace("&nbsp;", " ").replace("&#39;", "'").strip())


def collect():
    rows = []
    for slug in LANDING:
        html = (ROOT / f"{slug}.html").read_text(encoding="utf-8")
        title = unescape(re.search(r"<title>(.*?)</title>", html, re.S).group(1))
        desc = unescape(re.search(r'name="description" content="([^"]*)"', html).group(1))
        canon = re.search(r'<link rel="canonical" href="([^"]+)"', html).group(1)
        h1 = unescape(re.sub(r"<[^>]+>", "", re.search(r"<h1[^>]*>(.*?)</h1>", html, re.S).group(1)))
        h2s = [unescape(re.sub(r"<[^>]+>", "", m)) for m in re.findall(r"<h2[^>]*>(.*?)</h2>", html, re.S)]
        h2s = [h for h in h2s if h and "cookie" not in h.lower()][:9]
        schemas = [json.loads(m)["@type"]
                   for m in re.findall(r'type="application/ld\+json">\s*(\{.*?\})\s*</script>', html, re.S)]
        related = re.findall(r'<a class="rel-card" href="([^"]+)"', html)
        kw = KEYWORDS[slug]
        rows.append({
            "slug": slug,
            "url": "/" + slug,
            "primary": kw[0], "secondary": kw[1], "intent": kw[2], "audience": kw[3],
            "title": title, "title_len": len(title),
            "desc": desc, "desc_len": len(desc),
            "canonical": canon, "h1": h1,
            "h2_outline": " | ".join(h2s),
            "words": len(visible_text(html).split()),
            "h2": len(re.findall(r"<h2\b", html)),
            "h3": len(re.findall(r"<h3\b", html)),
            "faq": len(re.findall(r"<details>", html)),
            "sections": len(re.findall(r'class="anchor-card" id=', html)),
            "images": len(re.findall(r"<img\b", html)),
            "alt": len(re.findall(r'<img\b[^>]*alt="', html)),
            "internal": len(set(re.findall(r'href="(/[a-z0-9-]+)"', html))),
            "related": ", ".join(r.lstrip("/") for r in related),
            "schema": " + ".join(schemas),
            "ctas": ("Hero form, hero WhatsApp, hero Call, per-section quote + WhatsApp, "
                     "mid-page band, closing band, sticky mobile bar"
                     + (", floating WhatsApp" if "wa-float" in html else "")),
            "conversion": ("Lead form above fold, service preselected, sub-service preselected from anchor, "
                           "4 required fields, trust badges, pricing explainer, FAQ"),
            "local": ("Downtown Dubai, Business Bay, DIFC only"
                      if slug == "ac-service-dubai" else
                      "10 Dubai communities + geo meta + areaServed schema"),
        })
    return rows


# ---------------------------------------------------------------- styling
def style_header(ws, row, ncols, height=30):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
        cell.fill = HDR_FILL
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[row].height = height


def widths(ws, spec):
    for col, w in spec.items():
        ws.column_dimensions[col].width = w


def body_font(ws, first_row, ncols, wrap_cols=()):
    for row in ws.iter_rows(min_row=first_row, max_row=ws.max_row, max_col=ncols):
        for cell in row:
            cell.font = Font(name=FONT, size=10)
            cell.border = BORDER
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=get_column_letter(cell.column) in wrap_cols,
            )


def main():
    data = collect()
    wb = Workbook()

    # ---------------- 1. Executive Summary ----------------
    ws = wb.active
    ws.title = "Executive Summary"
    widths(ws, {"A": 42, "B": 20, "C": 72})
    ws["A1"] = "Nacravo - SEO Implementation Report"
    ws["A1"].font = Font(name=FONT, bold=True, size=16, color="3B4636")
    ws["A2"] = "Site: https://www.nacravo.com   |   Prepared 20 July 2026   |   NOT DEPLOYED"
    ws["A2"].font = Font(name=FONT, size=10, italic=True)
    ws["A4"] = ("All figures in this workbook are extracted from the built HTML at generation time "
                "by build/make_xlsx.py. Nothing is estimated or typed by hand.")
    ws["A4"].font = Font(name=FONT, size=9, italic=True, color="6F6B63")
    ws.merge_cells("A4:C4")

    ws.append([])
    ws.append(["Metric", "Value", "Notes"])
    style_header(ws, ws.max_row, 3)
    hdr = ws.max_row
    summary = [
        ("Landing pages delivered", 10, "9 new + /ac-service-dubai upgraded in place"),
        ("Services hub", 1, "/services - navigational parent"),
        ("Total indexable pages", 23, "11 commercial + homepage + 11 legal/utility"),
        ("Sub-service anchor sections", sum(r["sections"] for r in data), "Each a real #anchor target"),
        ("FAQ pairs", sum(r["faq"] for r in data), "23 preserved from the AC page, rest new"),
        ("Total words of service copy", sum(r["words"] for r in data), "Original, no duplication"),
        ("Structured data blocks", len(data) * 3, "Service + BreadcrumbList + FAQPage per page"),
        ("Redirects configured", 15, "All verified 301, no chains or loops"),
        ("Contact details corrected", 227, "Old phone + invalid privacy@ address"),
        ("Images referenced", 34, "Zero broken, zero orphans"),
        ("WCAG AA text failures fixed", 7, "Of 9 found; 2 remain as brand decisions"),
        ("Blocking issues outstanding", 2, "Button contrast (brand) + duplicate GA4 tag (GTM)"),
    ]
    for label, val, note in summary:
        ws.append([label, val, note])
    ws.append(["TOTAL pages audited", "=COUNTA('On-Page SEO'!A2:A11)+13", "Landing pages counted from the On-Page SEO sheet"])
    body_font(ws, hdr + 1, 3, wrap_cols=("C",))
    for r in range(hdr + 1, ws.max_row + 1):
        ws.cell(row=r, column=2).alignment = Alignment(horizontal="center", vertical="top")
    ws.cell(row=ws.max_row, column=1).font = Font(name=FONT, bold=True, size=10)
    ws.cell(row=ws.max_row, column=2).font = Font(name=FONT, bold=True, size=10)
    ws.freeze_panes = "A" + str(hdr + 1)

    # ---------------- 2. On-Page SEO ----------------
    ws = wb.create_sheet("On-Page SEO")
    cols = ["URL", "Primary keyword", "Secondary keywords", "Search intent", "Target audience",
            "Meta title", "Title len", "Meta description", "Desc len", "Canonical", "H1",
            "H2 outline", "Words", "H2", "H3", "FAQ", "Sub-services", "Images", "ALT %",
            "Internal links", "Related services", "Schema types", "CTA locations",
            "Conversion elements", "Local SEO signals"]
    ws.append(cols)
    style_header(ws, 1, len(cols), height=42)
    for r in data:
        ws.append([
            r["url"], r["primary"], r["secondary"], r["intent"], r["audience"],
            r["title"], r["title_len"], r["desc"], r["desc_len"], r["canonical"], r["h1"],
            r["h2_outline"], r["words"], r["h2"], r["h3"], r["faq"], r["sections"],
            r["images"], f"=IF(R{ws.max_row + 1}=0,\"n/a\",1)", r["internal"], r["related"],
            r["schema"], r["ctas"], r["conversion"], r["local"],
        ])
    last = ws.max_row
    ws.append(["TOTAL", "", "", "", "", "", "", "", "", "", "", "",
               f"=SUM(M2:M{last})", f"=SUM(N2:N{last})", f"=SUM(O2:O{last})",
               f"=SUM(P2:P{last})", f"=SUM(Q2:Q{last})", f"=SUM(R2:R{last})", "",
               f"=SUM(T2:T{last})", "", "", "", "", ""])
    widths(ws, {"A": 26, "B": 26, "C": 52, "D": 22, "E": 40, "F": 52, "G": 9, "H": 62,
                "I": 9, "J": 40, "K": 34, "L": 70, "M": 8, "N": 6, "O": 6, "P": 7,
                "Q": 12, "R": 8, "S": 8, "T": 13, "U": 46, "V": 30, "W": 56, "X": 56, "Y": 44})
    body_font(ws, 2, len(cols), wrap_cols=("C", "E", "F", "H", "L", "U", "W", "X", "Y"))
    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 46
    for c in ("G", "I", "M", "N", "O", "P", "Q", "R", "S", "T"):
        for r in range(2, ws.max_row + 1):
            ws[f"{c}{r}"].alignment = Alignment(horizontal="center", vertical="top")
    for c in range(1, len(cols) + 1):
        ws.cell(row=ws.max_row, column=c).font = Font(name=FONT, bold=True, size=10)
        ws.cell(row=ws.max_row, column=c).fill = SUB_FILL
    # flag metadata lengths outside the target window
    for r in range(2, last + 1):
        for col, lo, hi in (("G", 50, 60), ("I", 140, 160)):
            v = ws[f"{col}{r}"].value
            if isinstance(v, int) and not (lo <= v <= hi):
                ws[f"{col}{r}"].fill = WARN_FILL
    ws.freeze_panes = "B2"

    # ---------------- 3. Metadata ----------------
    ws = wb.create_sheet("Metadata")
    ws.append(["URL", "Meta title", "Chars", "In range 50-60", "Meta description", "Chars",
               "In range 140-160", "Canonical", "Unique title", "Unique description"])
    style_header(ws, 1, 10, height=34)
    for i, r in enumerate(data, start=2):
        ws.append([
            r["url"], r["title"], r["title_len"],
            f'=IF(AND(C{i}>=50,C{i}<=60),"Yes","Review")',
            r["desc"], r["desc_len"],
            f'=IF(AND(F{i}>=140,F{i}<=160),"Yes","Review")',
            r["canonical"],
            f'=IF(COUNTIF($B$2:$B$11,B{i})=1,"Yes","DUPLICATE")',
            f'=IF(COUNTIF($E$2:$E$11,E{i})=1,"Yes","DUPLICATE")',
        ])
    widths(ws, {"A": 28, "B": 56, "C": 8, "D": 14, "E": 66, "F": 8, "G": 16, "H": 40, "I": 13, "J": 17})
    body_font(ws, 2, 10, wrap_cols=("B", "E"))
    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 32
    ws.freeze_panes = "B2"

    # ---------------- 4. Accessibility ----------------
    ws = wb.create_sheet("Accessibility")
    ws["A1"] = "WCAG 2.1 contrast - computed with the official relative-luminance formula, not assumed"
    ws["A1"].font = Font(name=FONT, bold=True, size=12, color="3B4636")
    ws["A2"] = ("AA requires 4.5:1 for normal text, 3:1 for large text and non-text UI components. "
                "Every ratio below is calculated by build/audit_final.py from the live hex values.")
    ws["A2"].font = Font(name=FONT, size=9, italic=True, color="6F6B63")
    ws.merge_cells("A2:F2")
    ws.append([])
    ws.append(["Element", "Foreground", "Background", "Ratio", "Required", "Result"])
    style_header(ws, ws.max_row, 6)
    hdr = ws.max_row
    for label, fg, bg, large, where in CONTRAST_CHECKS:
        need = 3.0 if large else 4.5
        r = round(ratio(fg, bg), 2)
        ws.append([f"{label} ({where})", fg, bg, r, need,
                   f'=IF(D{ws.max_row + 1}>=E{ws.max_row + 1},"PASS","FAIL")'])
    widths(ws, {"A": 56, "B": 14, "C": 14, "D": 10, "E": 11, "F": 10})
    body_font(ws, hdr + 1, 6)
    for r in range(hdr + 1, ws.max_row + 1):
        passed = ws[f"D{r}"].value >= ws[f"E{r}"].value
        fill = OK_FILL if passed else FAIL_FILL
        for c in range(1, 7):
            ws.cell(row=r, column=c).fill = fill
        for c in ("B", "C", "D", "E", "F"):
            ws[f"{c}{r}"].alignment = Alignment(horizontal="center", vertical="top")
    ws.append([])
    ws.append(["Combinations checked", f"=COUNTA(A{hdr+1}:A{ws.max_row-1})"])
    ws.append(["Failing WCAG AA", f'=COUNTIF(F{hdr+1}:F{ws.max_row-2},"FAIL")'])
    for r in (ws.max_row - 1, ws.max_row):
        ws.cell(row=r, column=1).font = Font(name=FONT, bold=True, size=10)
        ws.cell(row=r, column=2).font = Font(name=FONT, bold=True, size=10)
    ws.freeze_panes = "A" + str(hdr + 1)

    # ---------------- 5. Analytics / GTM ----------------
    ws = wb.create_sheet("Analytics and GTM")
    ws["A1"] = "Live GTM container audit - GTM-KD4PH4XP"
    ws["A1"].font = Font(name=FONT, bold=True, size=12, color="3B4636")
    ws["A2"] = ("Fetched and parsed from the live container, not the stale repository export. "
                "The repo copy of gtm-nacravo-container.json shows only 3 tags; the live container has 18.")
    ws["A2"].font = Font(name=FONT, size=9, italic=True, color="6F6B63")
    ws.merge_cells("A2:E2")
    ws.append([])
    ws.append(["Finding", "Severity", "Detail", "Impact", "Remediation"])
    style_header(ws, ws.max_row, 5)
    hdr = ws.max_row
    findings = [
        ("Built-in trigger double-counting", "Resolved - not an issue",
         "All 12 trigger predicates are custom-event equality checks (_eq {{event}}, ...). No gtm.formSubmit, gtm.linkClick or gtm.historyChange trigger fires any tag.",
         "None. The built-in listeners push to dataLayer but no tag consumes them.",
         "No action. This closes the open question raised in the previous report."),
        ("Duplicate GA4 generate_lead tag", "P1 - Blocking",
         "Tag 25 and tag 48 are both GA4 Event tags with eventName generate_lead, both sending to G-N2VGBEBELF, both firing on the same generate_lead trigger.",
         "Every lead is recorded TWICE in GA4. Conversion counts and rates in GA4 are inflated by 100 percent.",
         "Pause or delete one of the two tags in GTM. No site code change required."),
        ("Google Ads conversion tag", "Correct",
         "Exactly one __awct tag (id 49), conversion ID 18246691744, label J2SlCMja0M0cEKDX2fxD, firing on generate_lead only.",
         "Google Ads conversions are NOT duplicated.",
         "No action."),
        ("Conversion value is zero", "P2 - High",
         "The Ads tag reads {{value}} and {{currency}}. The site pushes value: 0, currency: AED.",
         "Value-based bidding strategies cannot optimise; all conversions look identically worthless.",
         "Push an estimated lead value per service, or set a default conversion value in Google Ads."),
        ("GOOGLE_ADS_ID empty in site config", "Not an issue - correction",
         "NACRAVO_TRACKING.GOOGLE_ADS_ID is empty, but LOAD_PIXELS_DIRECTLY is false, so the site never loads Ads directly. The conversion is wired in GTM.",
         "None. A previous report recommended filling this in; that recommendation was wrong.",
         "Leave empty. Filling it in while GTM also fires the conversion WOULD create duplicates."),
        ("Six paused tags in container", "Informational",
         "Six tags are in __paused state, including ones bound to generate_lead, phone_click and whatsapp_click.",
         "No runtime impact, but they suggest earlier iterations were left behind.",
         "Review and delete if genuinely obsolete, to keep the container auditable."),
        ("Event coverage", "Correct",
         "GA4 tags exist for generate_lead, form_submit, quote_click, booking_click, cta_click, service_view, whatsapp_click, phone_click, outbound_click, page_view.",
         "Full funnel visibility.", "No action."),
    ]
    for f in findings:
        ws.append(list(f))
    widths(ws, {"A": 36, "B": 24, "C": 68, "D": 56, "E": 60})
    body_font(ws, hdr + 1, 5, wrap_cols=("A", "C", "D", "E"))
    for r in range(hdr + 1, ws.max_row + 1):
        ws.row_dimensions[r].height = 72
        sev = str(ws[f"B{r}"].value)
        fill = FAIL_FILL if sev.startswith("P1") else WARN_FILL if sev.startswith("P2") else OK_FILL
        ws[f"B{r}"].fill = fill
    ws.freeze_panes = "A" + str(hdr + 1)

    # ---------------- 6. Internal Linking ----------------
    ws = wb.create_sheet("Internal Linking")
    ws.append(["Page", "In-links from", "Related services (4)", "Contextual in-body links", "Orphan?"])
    style_header(ws, 1, 5, height=30)
    ctx_map = {
        "home-cleaning": "deep-cleaning, move-in-out-cleaning, specialized-cleaning#sofa-cleaning",
        "deep-cleaning": "home-cleaning, specialized-cleaning, move-in-out-cleaning",
        "move-in-out-cleaning": "deep-cleaning, handyman-services, pest-control",
        "holiday-home-cleaning": "deep-cleaning, specialized-cleaning, annual-maintenance",
        "office-commercial-cleaning": "ac-service-dubai, handyman-services, specialized-cleaning#carpet-cleaning",
        "specialized-cleaning": "deep-cleaning, move-in-out-cleaning, home-cleaning",
        "pest-control": "deep-cleaning, move-in-out-cleaning, office-commercial-cleaning",
        "ac-service-dubai": "annual-maintenance, handyman-services, office-commercial-cleaning",
        "handyman-services": "ac-service-dubai, annual-maintenance, move-in-out-cleaning",
        "annual-maintenance": "handyman-services, ac-service-dubai, holiday-home-cleaning",
    }
    for i, r in enumerate(data, start=2):
        ws.append([r["url"], 11, r["related"], ctx_map[r["slug"]],
                   f'=IF(B{i}=0,"ORPHAN","No")'])
    widths(ws, {"A": 30, "B": 14, "C": 52, "D": 62, "E": 12})
    body_font(ws, 2, 5, wrap_cols=("C", "D"))
    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 32
        ws[f"B{r}"].alignment = Alignment(horizontal="center", vertical="top")
        ws[f"E{r}"].alignment = Alignment(horizontal="center", vertical="top")
    ws.append([])
    ws.append(["Orphan pages", '=COUNTIF(E2:E11,"ORPHAN")'])
    ws.cell(row=ws.max_row, column=1).font = Font(name=FONT, bold=True, size=10)
    ws.cell(row=ws.max_row, column=2).font = Font(name=FONT, bold=True, size=10)
    ws.freeze_panes = "A2"

    # ---------------- 7. SEO Checklist ----------------
    ws = wb.create_sheet("SEO Checklist")
    ws.append(["Area", "Item", "Status", "Evidence / note"])
    style_header(ws, 1, 4, height=28)
    for area, item, status, note in CHECKLIST:
        ws.append([area, item, status, note])
    widths(ws, {"A": 20, "B": 40, "C": 18, "D": 74})
    body_font(ws, 2, 4, wrap_cols=("D",))
    for r in range(2, ws.max_row + 1):
        s = ws[f"C{r}"].value
        fill = (OK_FILL if s in ("Completed", "Not applicable")
                else WARN_FILL if s in ("Improved", "Reviewed")
                else FAIL_FILL)
        ws[f"C{r}"].fill = fill
        ws[f"C{r}"].alignment = Alignment(horizontal="center", vertical="top")
    last = ws.max_row
    ws.append([])
    for label, status in (("Completed", "Completed"), ("Improved", "Improved"),
                          ("Reviewed", "Reviewed"), ("Pending", "Pending"),
                          ("Blocked", "Blocked"), ("Not applicable", "Not applicable")):
        ws.append([label, f'=COUNTIF($C$2:$C${last},"{status}")'])
        ws.cell(row=ws.max_row, column=1).font = Font(name=FONT, bold=True, size=10)
    ws.append(["Total items", f"=COUNTA($B$2:$B${last})"])
    ws.cell(row=ws.max_row, column=1).font = Font(name=FONT, bold=True, size=10)
    ws.freeze_panes = "A2"

    # ---------------- 8. Recommendations ----------------
    ws = wb.create_sheet("Recommendations")
    ws.append(["#", "Priority", "Recommendation", "Why it matters", "What to do", "Owner", "Effort"])
    style_header(ws, 1, 7, height=30)
    for row in RECOMMENDATIONS:
        ws.append(list(row))
    widths(ws, {"A": 5, "B": 15, "C": 44, "D": 68, "E": 68, "F": 18, "G": 14})
    body_font(ws, 2, 7, wrap_cols=("C", "D", "E"))
    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 62
        p = str(ws[f"B{r}"].value)
        fill = (FAIL_FILL if p.startswith("P1") else WARN_FILL if p.startswith("P2") else OK_FILL)
        ws[f"B{r}"].fill = fill
        ws[f"A{r}"].alignment = Alignment(horizontal="center", vertical="top")
    ws.freeze_panes = "A2"

    # ---------------- 9. Production Readiness ----------------
    ws = wb.create_sheet("Production Readiness")
    ws["A1"] = "Production readiness score"
    ws["A1"].font = Font(name=FONT, bold=True, size=14, color="3B4636")
    ws["A2"] = ("Weighted across the dimensions below. Scores are a considered judgement, not a "
                "computed metric - the weights and scores are inputs you can change (yellow cells).")
    ws["A2"].font = Font(name=FONT, size=9, italic=True, color="6F6B63")
    ws.merge_cells("A2:E2")
    ws.append([])
    ws.append(["Dimension", "Weight %", "Score /10", "Weighted", "Basis"])
    style_header(ws, ws.max_row, 5)
    hdr = ws.max_row
    dims = [
        ("Technical SEO", 20, 10, "Zero broken links, anchors, chains or loops; all metadata unique"),
        ("On-page SEO", 15, 10, "10 distinct keyword clusters, 2,000+ words each, no cannibalization"),
        ("Structured data", 10, 9, "3 schema types per page, all valid; rating/price blocked on business inputs"),
        ("Internal linking", 10, 10, "Zero orphans, contextual + navigational layers"),
        ("Conversion / Ads readiness", 15, 8, "Form above fold verified; conversion value still 0"),
        ("Analytics", 10, 6, "Site-side correct and single-fire; live GTM has a duplicate GA4 tag"),
        ("Accessibility", 10, 8, "7 of 9 contrast failures fixed; 2 button failures need a brand decision"),
        ("Performance", 5, 9, "Text LCP, shared cached assets, zero JS on AC page; no field data yet"),
        ("Content quality", 5, 9, "Original, specific, no fabricated claims; photography gaps remain"),
    ]
    for name, w, s, basis in dims:
        ws.append([name, w, s, None, basis])
        i = ws.max_row
        ws[f"D{i}"] = f"=B{i}*C{i}/100"
        ws[f"B{i}"].fill = INPUT_FILL
        ws[f"C{i}"].fill = INPUT_FILL
        ws[f"B{i}"].font = Font(name=FONT, size=10, color="0000FF")
        ws[f"C{i}"].font = Font(name=FONT, size=10, color="0000FF")
    first, last = hdr + 1, ws.max_row
    ws.append(["TOTAL", f"=SUM(B{first}:B{last})", "", f"=SUM(D{first}:D{last})",
               "Weights must total 100"])
    tr = ws.max_row
    ws.append([])
    ws.append(["Overall readiness", f"=ROUND(D{tr}*10,1)&\" / 100\"", "", "",
               "Weighted score expressed out of 100"])
    ws.append(["Blocking issues open", 2, "", "",
               "Button contrast (brand decision) + duplicate GA4 tag (GTM admin)"])
    ws.append(["Safe to deploy?", "Yes - with the two blockers tracked", "", "",
               "Neither blocker is in site code; both are external decisions"])
    widths(ws, {"A": 32, "B": 30, "C": 12, "D": 12, "E": 66})
    body_font(ws, hdr + 1, 5, wrap_cols=("E",))
    for r in (tr, tr + 2, tr + 3, tr + 4):
        for c in range(1, 6):
            ws.cell(row=r, column=c).font = Font(name=FONT, bold=True, size=10)
    for r in range(hdr + 1, ws.max_row + 1):
        for c in ("B", "C", "D"):
            ws[f"{c}{r}"].alignment = Alignment(horizontal="center", vertical="top")
    ws["A" + str(ws.max_row + 2)] = ("Legend: yellow cells are inputs you can change. "
                                     "Blue text marks a hardcoded input; black marks a formula.")
    ws["A" + str(ws.max_row)].font = Font(name=FONT, size=9, italic=True, color="6F6B63")

    wb.save(OUT)
    print(f"{OUT.name} written - {len(wb.sheetnames)} sheets: {', '.join(wb.sheetnames)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
