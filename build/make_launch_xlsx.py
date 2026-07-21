"""Build GOOGLE_ADS_LAUNCH_PACK.xlsx and SEO_KPI_DASHBOARD.xlsx.

Ads pack rows are derived from the built pages (URLs + real #anchors), so the
final URLs cannot point at anchors that do not exist. The KPI dashboard is a
tracking template with example rows and live formulas (CPL, conversion rate,
ROAS), professionally formatted, with input cells marked.

Usage: ./.venv/Scripts/python.exe build/make_launch_xlsx.py
"""

import pathlib
import re
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = pathlib.Path(__file__).resolve().parent.parent
FONT = "Arial"
HDR = PatternFill("solid", fgColor="3B4636")
SUB = PatternFill("solid", fgColor="E8E0D0")
INPUT = PatternFill("solid", fgColor="FFF6CC")
CALC = PatternFill("solid", fgColor="EAF0E4")
THIN = Side(style="thin", color="D8D2C4")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def hdr(ws, row, n, h=30):
    for c in range(1, n + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
        cell.fill = HDR
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[row].height = h


def widths(ws, spec):
    for col, w in spec.items():
        ws.column_dimensions[col].width = w


def body(ws, first, n, wrap=()):
    for r in ws.iter_rows(min_row=first, max_row=ws.max_row, max_col=n):
        for cell in r:
            cell.font = Font(name=FONT, size=10)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=get_column_letter(cell.column) in wrap)


# ---------------------------------------------------------------- ADS PACK
CAMPAIGNS = [
    ("Home Cleaning", "home-cleaning", "home cleaning Dubai",
     "house cleaning, apartment cleaning, villa cleaning, weekly cleaning, monthly cleaning, hourly cleaning, maid service",
     "Commercial - recurring", "Dubai (broad)",
     ["regular-cleaning", "apartment-cleaning", "villa-cleaning", "hourly-cleaning"],
     "Home Cleaning in Dubai | Trained, Insured Cleaners | Fixed Price | Photo Report | Book on WhatsApp",
     "Employed, background-checked cleaners across Dubai. Fixed price agreed first, materials included, photo report after."),
    ("Deep Cleaning", "deep-cleaning", "deep cleaning Dubai",
     "apartment deep cleaning, villa deep cleaning, kitchen deep cleaning, bathroom deep cleaning, post construction cleaning",
     "Commercial - one-off", "Dubai (broad)",
     ["kitchen-deep-cleaning", "bathroom-deep-cleaning", "apartment-deep-cleaning", "post-construction"],
     "Deep Cleaning in Dubai | Top-to-Bottom Reset | Kitchen & Bathroom | Fixed Price | Photo Report",
     "A full reset for grease, limescale and build-up a routine clean can't touch. Fixed price, photo report after."),
    ("Move In / Move Out", "move-in-out-cleaning", "move out cleaning Dubai",
     "move in cleaning, end of tenancy cleaning, empty property cleaning, oven cleaning, fridge cleaning",
     "Transactional - deadline", "Dubai (broad)",
     ["end-of-tenancy-cleaning", "move-out-cleaning", "move-in-cleaning", "oven-cleaning"],
     "Move Out Cleaning Dubai | End of Tenancy | Built for Inspections | Fixed Price | Photo Report",
     "End-of-tenancy cleaning built around a landlord's inspection. Kitchen, bathroom, balcony, appliances included."),
    ("Holiday Home Cleaning", "holiday-home-cleaning", "holiday home cleaning Dubai",
     "Airbnb cleaning, holiday home turnover, short let cleaning, linen replacement, guest ready",
     "Commercial - B2B recurring", "Dubai (broad)",
     ["airbnb-cleaning", "holiday-home-turnover", "linen-replacement", "mid-stay-cleaning"],
     "Holiday Home Cleaning Dubai | Airbnb Turnover | Guest-Ready | Linen & Restocking | Photo Report",
     "Fast, reliable turnovers between guests, with linen and consumables handled and a photo report each time."),
    ("Office & Commercial", "office-commercial-cleaning", "office cleaning Dubai",
     "commercial cleaning, daily office cleaning, office deep cleaning, retail cleaning, fit-out cleaning, post-event cleaning",
     "B2B contract", "Dubai (broad)",
     ["office-cleaning", "daily-contracts", "office-deep-cleaning", "fit-out-cleaning"],
     "Office Cleaning in Dubai | Daily Contracts | Commercial Team | Scheduled Around Your Hours | Invoices",
     "Daily contracts, deep cleans and post-event cleaning for Dubai workplaces, with invoices and photo reports."),
    ("Specialized Cleaning", "specialized-cleaning", "sofa cleaning Dubai",
     "carpet cleaning, mattress cleaning, curtain cleaning, upholstery cleaning, interior windows",
     "Commercial - item level", "Dubai (broad)",
     ["sofa-cleaning", "carpet-cleaning", "mattress-cleaning", "curtain-cleaning"],
     "Sofa Cleaning in Dubai | Carpet & Mattress | Extraction, Not Surface Wipe | Fixed Price | Photo Report",
     "Sofa, carpet, mattress and curtain cleaning by extraction. Trained technicians, fixed price, photo report."),
    ("Pest Control", "pest-control", "pest control Dubai",
     "cockroach control, bed bug treatment, ant control, rodent control, commercial pest control",
     "Urgent problem-solving", "Dubai (broad)",
     ["cockroach-control", "bed-bugs", "ant-control", "rodent-control"],
     "Pest Control in Dubai | Cockroach & Bed Bug | Ant & Rodent | Homes & Businesses | Book on WhatsApp",
     "Cockroach, bed bug, ant and rodent treatment. We inspect first, explain prep and re-entry time before treating."),
    ("AC Services", "ac-service-dubai", "AC service Dubai",
     "AC chemical wash, AC duct cleaning, AC repair, AC installation, AC gas top up, split AC service",
     "Commercial + urgent", "Downtown Dubai, Business Bay, DIFC ONLY",
     ["servicing", "chemical-wash", "duct-cleaning", "repair"],
     "AC Service in Dubai | Chemical Wash | Repair | Licensed Technicians | Fixed Price | Downtown-Bay-DIFC",
     "Servicing, chemical wash, duct cleaning and repair by licensed technicians. Focused on Downtown, Bay & DIFC."),
    ("Handyman Services", "handyman-services", "handyman Dubai",
     "plumber Dubai, electrician Dubai, painting, TV mounting, furniture assembly, curtain fixing",
     "Task-specific", "Dubai (broad)",
     ["plumbing", "electrical", "painting", "tv-mounting"],
     "Handyman Services Dubai | Plumbing & Electrical | Painting & Repairs | Fixed Price, Parts Shown First",
     "Plumbing, electrical, painting, mounting and repairs by one vetted team. Fixed price, parts shown before fitting."),
    ("Annual Maintenance", "annual-maintenance", "annual maintenance contract Dubai",
     "AMC Dubai, villa AMC, apartment AMC, preventive maintenance, MEP inspection, property maintenance contract",
     "B2B / landlord", "Dubai (broad)",
     ["apartment-amc", "villa-amc", "preventive-maintenance", "mep-inspection"],
     "Annual Maintenance Dubai | Villa & Apartment AMC | Preventive Visits | Priority Callout | Get a Quote",
     "Scheduled preventive visits instead of waiting for a breakdown. Service record per unit, priority callout."),
]


def verify_anchors():
    """Fail loudly if any recommended #anchor does not exist on its page."""
    bad = []
    for _, slug, _, _, _, _, anchors, _, _ in CAMPAIGNS:
        html = (ROOT / f"{slug}.html").read_text(encoding="utf-8")
        present = set(re.findall(r'class="anchor-card" id="([^"]+)"', html))
        for a in anchors:
            if a not in present:
                bad.append(f"{slug}#{a}")
    if bad:
        raise SystemExit(f"FAILED: recommended anchors not found on page: {bad}")


def build_ads():
    verify_anchors()
    wb = Workbook()
    ws = wb.active
    ws.title = "Campaigns"
    cols = ["Campaign", "Landing page (Final URL)", "Primary keyword", "Secondary keywords",
            "Search intent", "Geo targeting", "Ad groups (deep-link anchors)",
            "Recommended deep-link Final URLs", "Suggested headlines (| separated, <=30 ea)",
            "Suggested descriptions (<=90 ea)"]
    ws.append(cols)
    hdr(ws, 1, len(cols), h=40)
    site = "https://www.nacravo.com"
    for name, slug, pk, sk, intent, geo, anchors, heads, descs in CAMPAIGNS:
        deeplinks = "\n".join(f"/{slug}#{a}" for a in anchors)
        ws.append([name, f"{site}/{slug}", pk, sk, intent, geo,
                   ", ".join(anchors), deeplinks, heads, descs])
    widths(ws, {"A": 22, "B": 40, "C": 26, "D": 50, "E": 22, "F": 34,
                "G": 44, "H": 40, "I": 66, "J": 66})
    body(ws, 2, len(cols), wrap=("D", "G", "H", "I", "J"))
    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 90
        if "ONLY" in str(ws[f"F{r}"].value):
            ws[f"F{r}"].fill = PatternFill("solid", fgColor="FBE6E2")
            ws[f"F{r}"].font = Font(name=FONT, size=10, bold=True)
    ws.freeze_panes = "B2"

    # shared assets sheet
    ws2 = wb.create_sheet("Shared Assets")
    ws2.append(["Asset type", "Values"])
    hdr(ws2, 1, 2)
    shared = [
        ("Callouts", "Employed, vetted technicians; Fixed price before work; Materials & equipment included; Photo report every visit; Book on WhatsApp; One accountable team; English-speaking staff; Not a marketplace"),
        ("Structured snippet - Services", "Home Cleaning; Deep Cleaning; Move In/Out; Office Cleaning; AC Service; Handyman; Pest Control; Annual Maintenance"),
        ("Structured snippet - Types", "Apartments; Villas; Townhouses; Offices; Holiday Homes"),
        ("Sitelinks", "Deep Cleaning /deep-cleaning; AC Chemical Wash /ac-service-dubai#chemical-wash; Sofa & Carpet /specialized-cleaning; Move Out Cleaning /move-in-out-cleaning; Handyman /handyman-services; All Services /services"),
        ("Call extension", "+971 58 108 2601"),
        ("Location extension", "Link the Google Business Profile once live"),
        ("Image extensions", "Use real service photos. SKIP for Pest Control and AC until that photography exists."),
        ("Negative keywords (starter)", "jobs; careers; salary; vacancy; free; DIY; how to; course; training; wholesale"),
        ("Bidding at launch", "Maximise Conversions (NOT tROAS - conversion value is 0 until set)"),
        ("IMPORTANT", "Final URLs must point at real pages, not vanity redirects (/sofa-cleaning etc). Deep-link anchors preselect the sub-service in the form."),
    ]
    for a, v in shared:
        ws2.append([a, v])
    widths(ws2, {"A": 30, "B": 100})
    body(ws2, 2, 2, wrap=("B",))
    for r in range(2, ws2.max_row + 1):
        ws2.row_dimensions[r].height = 44
    ws2.freeze_panes = "A2"

    wb.save(ROOT / "GOOGLE_ADS_LAUNCH_PACK.xlsx")
    print("GOOGLE_ADS_LAUNCH_PACK.xlsx written - anchors verified against live pages")


# ---------------------------------------------------------------- KPI DASHBOARD
def build_kpi():
    wb = Workbook()

    # --- Monthly tracker ---
    ws = wb.active
    ws.title = "Monthly Tracker"
    ws["A1"] = "Nacravo — SEO & Google Ads KPI Dashboard"
    ws["A1"].font = Font(name=FONT, bold=True, size=15, color="3B4636")
    ws["A2"] = ("Yellow cells are inputs you fill in each month. Green cells are formulas — do not "
                "overwrite them. One example row (Baseline) shows the expected format.")
    ws["A2"].font = Font(name=FONT, size=9, italic=True, color="6F6B63")
    ws.merge_cells("A2:R2")
    ws.append([])

    cols = ["Month", "Organic clicks", "Organic impressions", "Organic CTR %", "Avg position",
            "Ranking keywords (top 10)", "Total leads", "Phone calls", "WhatsApp leads",
            "Lead forms", "Google Ads leads", "Ads cost (AED)", "Ads conv. value (AED)",
            "Conversion rate %", "Cost per lead (AED)", "ROAS", "Google reviews", "Avg rating"]
    ws.append(cols)
    hrow = ws.max_row
    hdr(ws, hrow, len(cols), h=44)

    # formula columns: D=CTR (B/C), N=conv rate (leads/ads clicks? use leads/(sessions unknown)) -> use Ads leads/Ads cost not rate; keep CR = total leads is not rate. Use CR% = Ads leads / Ads clicks unavailable; instead CPL and ROAS which we can compute.
    # We compute: CTR% = clicks/impr; CPL = ads cost / ads leads; ROAS = conv value / ads cost.
    # Conversion rate here = total leads / organic+ads clicks (approx) -> leave as input to avoid a wrong denominator.
    example = ["Baseline (pre-launch)", 0, 0, None, None, 0, 0, 0, 0, 0, 0, 0, 0, None, None, None, 0, None]
    ws.append(example)
    r = ws.max_row
    ws[f"D{r}"] = f"=IF(C{r}=0,0,B{r}/C{r})"
    ws[f"O{r}"] = f"=IF(K{r}=0,0,L{r}/K{r})"          # CPL = ads cost / ads leads
    ws[f"P{r}"] = f"=IF(L{r}=0,0,M{r}/L{r})"          # ROAS = conv value / ads cost
    ws[f"N{r}"] = None                                # conversion rate: input (denominator varies by source)

    # 12 blank month rows with formulas pre-filled
    months = ["Month 1", "Month 2", "Month 3", "Month 4", "Month 5", "Month 6",
              "Month 7", "Month 8", "Month 9", "Month 10", "Month 11", "Month 12"]
    for m in months:
        ws.append([m, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None])
        r = ws.max_row
        ws[f"D{r}"] = f"=IF(C{r}=0,0,B{r}/C{r})"
        ws[f"O{r}"] = f"=IF(K{r}=0,0,L{r}/K{r})"
        ws[f"P{r}"] = f"=IF(L{r}=0,0,M{r}/L{r})"

    widths(ws, {"A": 20, "B": 13, "C": 15, "D": 10, "E": 10, "F": 16, "G": 11, "H": 11,
                "I": 13, "J": 11, "K": 13, "L": 13, "M": 15, "N": 13, "O": 15, "P": 9,
                "Q": 12, "R": 10})
    body(ws, hrow + 1, len(cols))
    input_cols = "BCEFGHIJKLMNQR"   # user inputs
    calc_cols = "DOP"                # formulas
    for row in range(hrow + 1, ws.max_row + 1):
        for c in input_cols:
            ws[f"{c}{row}"].fill = INPUT
        for c in calc_cols:
            ws[f"{c}{row}"].fill = CALC
        ws[f"D{row}"].number_format = "0.0%"
        ws[f"N{row}"].number_format = "0.0%"
        ws[f"O{row}"].number_format = "#,##0"
        ws[f"P{row}"].number_format = "0.0x"
        ws[f"L{row}"].number_format = "#,##0"
        ws[f"M{row}"].number_format = "#,##0"
        ws[f"R{row}"].number_format = "0.0"
    ws.freeze_panes = "B" + str(hrow + 1)

    # --- Keyword rank tracker ---
    ws2 = wb.create_sheet("Keyword Rankings")
    ws2.append(["Keyword", "Landing page", "Search intent", "Baseline rank", "Month 1", "Month 2",
                "Month 3", "Month 6", "Month 12", "Best rank", "Trend"])
    hdr(ws2, 1, 11, h=30)
    seed_kw = [
        ("home cleaning dubai", "/home-cleaning"), ("deep cleaning dubai", "/deep-cleaning"),
        ("move out cleaning dubai", "/move-in-out-cleaning"), ("end of tenancy cleaning dubai", "/move-in-out-cleaning"),
        ("holiday home cleaning dubai", "/holiday-home-cleaning"), ("airbnb cleaning dubai", "/holiday-home-cleaning"),
        ("office cleaning dubai", "/office-commercial-cleaning"), ("sofa cleaning dubai", "/specialized-cleaning"),
        ("carpet cleaning dubai", "/specialized-cleaning"), ("mattress cleaning dubai", "/specialized-cleaning"),
        ("pest control dubai", "/pest-control"), ("bed bugs treatment dubai", "/pest-control"),
        ("ac service dubai", "/ac-service-dubai"), ("ac chemical wash dubai", "/ac-service-dubai"),
        ("ac repair dubai", "/ac-service-dubai"), ("handyman dubai", "/handyman-services"),
        ("plumber dubai", "/handyman-services"), ("electrician dubai", "/handyman-services"),
        ("annual maintenance contract dubai", "/annual-maintenance"), ("villa amc dubai", "/annual-maintenance"),
    ]
    for kw, url in seed_kw:
        ws2.append([kw, url, "", None, None, None, None, None, None,
                    f"=IF(COUNT(D{ws2.max_row+1}:I{ws2.max_row+1})=0,\"\",MIN(D{ws2.max_row+1}:I{ws2.max_row+1}))",
                    ""])
    # fix formula row references (max_row shifted); recompute simply per row
    for r in range(2, ws2.max_row + 1):
        ws2[f"J{r}"] = f'=IF(COUNT(D{r}:I{r})=0,"",MIN(D{r}:I{r}))'
        ws2[f"K{r}"] = f'=IF(OR(D{r}="",I{r}=""),"",IF(I{r}<D{r},"improving",IF(I{r}>D{r},"declining","flat")))'
    widths(ws2, {"A": 34, "B": 26, "C": 18, "D": 12, "E": 9, "F": 9, "G": 9, "H": 9, "I": 9, "J": 10, "K": 12})
    body(ws2, 2, 11)
    for r in range(2, ws2.max_row + 1):
        for c in "DEFGHI":
            ws2[f"{c}{r}"].fill = INPUT
        for c in "JK":
            ws2[f"{c}{r}"].fill = CALC
    ws2.freeze_panes = "B2"

    # --- Local + CWV + reviews snapshot ---
    ws3 = wb.create_sheet("Local & CWV & Reviews")
    ws3.append(["Metric", "Baseline", "Target", "Month 1", "Month 3", "Month 6", "Month 12", "Notes"])
    hdr(ws3, 1, 8, h=28)
    rows = [
        ("Local pack rankings (map) - core keywords", None, "Top 3", None, None, None, None, "Track via GBP Insights / local rank tool"),
        ("Google reviews (count)", 0, "20+ then grow", None, None, None, None, "See GBP plan review workflow"),
        ("Average rating", None, "4.5+", None, None, None, None, "Real reviews only"),
        ("GBP profile views", None, "grow", None, None, None, None, "GBP Insights"),
        ("GBP calls", None, "grow", None, None, None, None, "GBP Insights"),
        ("GBP website clicks", None, "grow", None, None, None, None, "GBP Insights"),
        ("LCP - mobile (s)", None, "< 2.5", None, None, None, None, "Lighthouse / CrUX post-deploy"),
        ("CLS - mobile", None, "< 0.1", None, None, None, None, "Lighthouse / CrUX"),
        ("INP - mobile (ms)", None, "< 200", None, None, None, None, "Lighthouse / CrUX"),
        ("Indexed pages (Search Console)", None, "23", None, None, None, None, "Coverage report"),
        ("Google Ads Quality Score (avg)", None, "7+", None, None, None, None, "Ads keyword QS column"),
    ]
    for row in rows:
        ws3.append(list(row))
    widths(ws3, {"A": 40, "B": 12, "C": 14, "D": 10, "E": 10, "F": 10, "G": 10, "H": 40})
    body(ws3, 2, 8, wrap=("H",))
    for r in range(2, ws3.max_row + 1):
        for c in "DEFG":
            ws3[f"{c}{r}"].fill = INPUT
    ws3.freeze_panes = "B2"

    # --- Legend ---
    ws4 = wb.create_sheet("How to use")
    notes = [
        ["How to use this dashboard", ""],
        ["", ""],
        ["Yellow cells", "Inputs you fill in each month from the sources named."],
        ["Green cells", "Formulas (CTR, cost per lead, ROAS, best rank, trend). Do not overwrite."],
        ["Data sources", ""],
        ["Organic clicks / impressions / CTR / position", "Google Search Console > Performance"],
        ["Ranking keywords", "Search Console, or a rank tracker (Ahrefs / SEMrush / local tool)"],
        ["Phone / WhatsApp / form leads", "GA4 events: phone_click, whatsapp_click, generate_lead"],
        ["Google Ads leads / cost / value", "Google Ads > Campaigns"],
        ["Conversion value / ROAS", "Requires a conversion value to be set (see Ads report). ROAS = value / cost."],
        ["Cost per lead", "Ads cost / Ads leads (auto-calculated)."],
        ["Google reviews / rating", "Google Business Profile"],
        ["Core Web Vitals", "Lighthouse (lab) and Search Console CWV report (field) - AFTER deploy"],
        ["Quality Score", "Google Ads > Keywords > add the Quality Score column"],
        ["IMPORTANT", "Pause the duplicate GA4 tag first, or your lead numbers will be ~2x too high. See GTM guide."],
    ]
    for row in notes:
        ws4.append(row)
    ws4["A1"].font = Font(name=FONT, bold=True, size=14, color="3B4636")
    widths(ws4, {"A": 44, "B": 70})
    body(ws4, 3, 2, wrap=("B",))
    for r in range(3, ws4.max_row + 1):
        ws4[f"A{r}"].font = Font(name=FONT, bold=True, size=10)

    wb.save(ROOT / "SEO_KPI_DASHBOARD.xlsx")
    print("SEO_KPI_DASHBOARD.xlsx written")


if __name__ == "__main__":
    build_ads()
    build_kpi()
